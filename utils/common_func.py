import json

from pathlib import Path
from openpyxl import load_workbook
from typing import Tuple, Optional, Iterable
from loguru import logger


def get_field_type(field_type: str, mapping_fields: dict) -> str:
    return mapping_fields[field_type.strip()]


def save_dictionaries(
    schema_dictionaries: dict, schema_dictionaries_path: Path
):
    for file_name, schema_dict_data in schema_dictionaries.items():
        if Path.is_file(Path(schema_dictionaries_path, file_name)):
            logger.warning("File {} exists and will not be saved.".format(file_name))
            continue
        save_schema_to_json(schema_dict_data, schema_dictionaries_path / file_name)


def save_schema_to_files(
    schema: dict,
    schema_path: Path,
) -> None:
    if "dictionaries" in schema.keys():
        save_dictionaries(schema["dictionaries"], schema_path)

    if "vacancy_fields" in schema.keys():
        save_schema_to_json(
            schema["vacancy_fields"],
            schema_path / schema["vacancy_fields_file_name"],
        )

    if "vacancy_request" in schema.keys():
        save_schema_to_json(
            schema["vacancy_request"],
            schema_path / schema["vacancy_request_file_name"],
        )

    if "questionary" in schema.keys():
        save_schema_to_json(
            schema["questionary"],
            schema_path / schema["questionary_fields_file_name"],
        )

    if "template_form" in schema.keys():
        save_schema_to_json(schema["template_form"], schema_path / "template.json"),


def get_cells(
    xlsx_filename,
    number_of_columns: int = 10,
) -> Iterable[
    Tuple[str, str, Optional[str], Optional[str], str, str, str, str, str, str]
]:
    """
    :param: active_worksheet
    :return: генератор возвращает кортеж. Каждый кортеж, соответствует строке таблицы.
    Если у обрабатываемой строки первая ячейка пустая - цикл прерывается. Обработаются только те ячейки,
    которые были прочитаны до прерывания цикла.
    """
    wb = load_workbook(filename=xlsx_filename, read_only=True)
    active_worksheet = wb.active

    for row in active_worksheet.iter_rows(min_row=2, max_col=number_of_columns):
        if row[0].value is None:
            break

        current_row_cells: list = []

        for cell in row:
            current_row_cells.append(cell.value)

        yield tuple(current_row_cells)


def save_schema_to_json(divisions: dict, filename: Path) -> None:
    with open(filename, "w", encoding="utf-8") as file:
        json.dump(divisions, file, ensure_ascii=False, indent=2)


def load_json(filename: Path) -> dict:
    with open(filename, "r", encoding="utf-8") as file:
        file_content = json.load(file)
        return file_content


def order_redefine(schema: dict, current_order: int = 1):
    for field_name, field_attribute in schema.items():
        sub_fields = [
            "compensation",
            "complex",
        ]
        fields = field_attribute.get("fields")
        field_attribute["order"] = current_order
        current_order += 1

        if field_attribute["type"] in sub_fields and fields:
            order = len(field_attribute["fields"])
            order_redefine(field_attribute["fields"], current_order)
            current_order += order  # Т.к. 2 внутренних поля типа compensation

    return schema


def fill_template_form(
    vacancy_request_data: dict,
    org_name: str,
    schema_path: Path,
    region_field: bool = False,
) -> dict:
    template_form = {
        "params": {"name": ""},
    }

    template_form["params"]["name"] = org_name

    if region_field:
        template_form["region"] = True

    if "vacancy_request_file_name" in vacancy_request_data.keys():
        vacancy_request_file_name = vacancy_request_data["vacancy_request_file_name"]
        template_form.setdefault("vacancy_request", [])
        template_form["vacancy_request"].append(vacancy_request_file_name)

    if (
        "vacancy_request" in vacancy_request_data
        and "account_division"
        in vacancy_request_data["vacancy_request"]["schema"].keys()
    ):
        template_form.setdefault("division", f"division_{schema_path}.json")

    if "vacancy_fields" in vacancy_request_data:
        vacancy_field_file_name = vacancy_request_data["vacancy_fields_file_name"]
        template_form.setdefault("vacancy_field", vacancy_field_file_name)

    if "dictionaries" in vacancy_request_data:
        template_form.setdefault("dictionary", [])
        for dict_file_name in vacancy_request_data["dictionaries"]:
            template_form["dictionary"].append(dict_file_name)

    if "questionary" in vacancy_request_data:
        template_form.setdefault(
            "questionary", f"schema_applicant_questionary_{schema_path}.json"
        )

    return template_form
